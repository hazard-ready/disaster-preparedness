import openpyxl  # library to read .xlsx format


def checkHTMLTagClosures(row, rowCount):
    tags_to_check = ["ol", "ul", "li", "a", "b", "i"]
    mismatches = {}
    for key in row.keys():
        if key == 'text' or key.startswith('text-'):
            for tag in tags_to_check:
                tag_opening_no_attr = "<" + tag + ">"
                tag_opening_with_attr = "<" + tag + " "
                tag_closing = "</" + tag + ">"
                tag_openings = row[key].count(
                    tag_opening_no_attr) + row[key].count(tag_opening_with_attr)
                tag_closings = row[key].count(tag_closing)
                if tag_openings > tag_closings:
                    print("WARNING: In row", str(rowCount), key, "has", str(
                        tag_openings), "opening <" + tag + "> tag[s] but only", str(tag_closings), "closing tag[s].")


def allRequiredFieldsPresent(optionalFields, row, rowCount):
    if any(a != '' for a in row.values()):  # if the entire row is not empty
        blanks = []
        for key in row.keys():
            if (key not in optionalFields) and key != '' and (row[key] == ''):
                blanks.append(key)
        if blanks == []:
            return True
        else:
            print("Unable to process row", rowCount, "with content:")
            print(row)
            if len(blanks) > 1:
                print("Because required fields", blanks, "are empty.")
            else:
                print("Because required field", "'" +
                      blanks[0] + "'", "is empty.")
            return False
    else:  # the entire row is blank
        print("Skipping empty row", rowCount)
        return False


def includeTranslatedFields(row, columnName, fieldName, kwargs):
    for translatedField in (k for k in row if k.startswith(columnName + '-')):
        if translatedField is not None and row[translatedField] != '':
            translatedColumn=fieldName + '_' + translatedField.split('-')[1]
            kwargs[translatedColumn]=row[translatedField]

    return kwargs


# drop-in replacement for built-in csv.DictReader() function with .xlsx files
# originally from https://gist.github.com/mdellavo/853413
# then heavily adapted first to make it work, then to simplify, and finally with suggestions from later commenters on that gist
def XLSXDictReader(fileName, sheetName=None):
    book=openpyxl.reader.excel.load_workbook(fileName)
    # if there's no sheet name specified, try to get the active sheet.  This will work reliably for workbooks with only one sheet; unpredictably if there are multiple worksheets present.
    if sheetName is None:
        sheet=book.active
    elif sheetName not in book.sheetnames:
        print(sheetName, "not found in", fileName)
        exit()
    else:
        sheet=book[sheetName]

    rows=sheet.max_row + 1
    cols=sheet.max_column + 1

    def cleanValue(s):
        if s == None:
            return ''
        else:
            return str(s).strip()

    def item(i, j):
        return (
            cleanValue(sheet.cell(row=1, column=j).value),
            cleanValue(sheet.cell(row=i, column=j).value)
        )

    return (dict(item(i, j) for j in range(1, cols)) for i in range(2, rows))


def runLoader(config):
    overwriteAll=False

    newSnuggets=XLSXDictReader(config['file'])
    # row 1 consists of field names, so row 2 is the first data row. We'll increment this before first referencing it.
    rowCount=1
    for row in newSnuggets:
        rowCount += 1
        if allRequiredFieldsPresent(config['optional'], row, rowCount):
            checkHTMLTagClosures(row, rowCount)
            overwriteAll=config['processRow'](row, overwriteAll)

    return rowCount
