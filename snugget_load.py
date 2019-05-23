import openpyxl # library to read .xlsx format
import os

from django.core.files import File
from django.contrib.contenttypes.models import ContentType
from disasterinfosite.models import *

currentPath = os.path.dirname(__file__)
appName = "disasterinfosite"
appDir = os.path.join(currentPath, "disasterinfosite")
dataDir = os.path.join(appDir, "data")
snuggetFile = os.path.join(dataDir, "snuggets.xlsx")
slideshowFilename = "slideshow.xlsx" # there can be multiple of these files in different locations, so the full path is assembled in addSlideshow()

requiredFields = ['shapefile', 'section', 'subsection']
# all other fields in snuggetFile are required. The empty string is to deal with Excel's charming habit of putting a blank column after all data in a CSV.
optionalFields = ['heading', 'intensity', 'lookup_value', 'txt_location', 'pop_out_image', 'pop_out_link','pop_alt_txt', 'pop_out_txt', 'pop_out_video', 'intensity_txt', 'text', 'image_slideshow_folder', 'video', '']

defaults = {
  "intensity": None,
  "pop_out_video": None,
  "txt_location": 0
}

def run():
  overwriteAll = False

  newSnuggets = XLSXDictReader(snuggetFile)
  rowCount = 1 # row 1 consists of field names, so row 2 is the first data row. We'll increment this before first referencing it.
  for row in newSnuggets:
    rowCount += 1
    if allRequiredFieldsPresent(row, rowCount):
      checkHTMLTagClosures(row, rowCount)
      overwriteAll = processRow(row, overwriteAll)

  print("Snugget load complete. Processed", rowCount, "rows in", snuggetFile)


def allRequiredFieldsPresent(row, rowCount):
  if any(a != '' for a in row.values()): # if the entire row is not empty
    blanks = []
    for key in row.keys():
      if (key not in optionalFields) and (row[key] == ''):
        blanks.append(key)
    if blanks == []:
      return True
    else:
      print("Unable to process row", rowCount, "with content:")
      print(row)
      if len(blanks) > 1:
        print("Because required fields", blanks, "are empty.")
      else:
        print("Because required field", "'" + blanks[0] + "'", "is empty.")
      return False
  else: # the entire row is blank
    print("Skipping empty row", rowCount)
    return False


def checkHTMLTagClosures(row, rowCount):
  tags_to_check = ["ol", "ul", "li", "a", "b"]
  mismatches = {}
  for key in row.keys():
    if key == 'text' or key.startswith('text-'):
      for tag in tags_to_check:
        tag_opening_no_attr = "<" + tag + ">"
        tag_opening_with_attr = "<" + tag + " "
        tag_closing = "</" + tag + ">"
        tag_openings = row[key].count(tag_opening_no_attr) + row[key].count(tag_opening_with_attr)
        tag_closings = row[key].count(tag_closing)
        if tag_openings > tag_closings:
          print("WARNING: In row", str(rowCount), key, "has", str(tag_openings), "opening <" + tag + "> tag[s] but only", str(tag_closings), "closing tag[s].")


def setDefaults(row):
  for key, value in defaults.items():
    if row[key] == '':
      row[key] = value


def processRow(row, overwriteAll):
  shapefile = getShapefileClass(row)
  filterColumn = row["shapefile"] + "_filter"
  section, created = SnuggetSection.objects.get_or_create(name=row["section"])

  setDefaults(row)

  order = row["txt_location"]

  if created:
    print("Created a new snugget section: ", row["section"])

  # check if a snugget for this data already exists
  # if we have a lookup value then deal with this value specifically:
  if row["lookup_value"] is not '':  # if it is blank, we'll treat it as matching all existing values
    filterVal = row["lookup_value"]
    oldSnugget = checkForSnugget(shapefile, section, order, filterColumn, filterVal)
    overwriteAll = askUserAboutOverwriting(row, oldSnugget, overwriteAll)
    processSnugget(row, shapefile, section, order, filterColumn, filterVal)
  else:
    filterVals = findAllFilterVals(shapefile)
    oldSnuggets = []
    for filterVal in filterVals:
      oldSnugget = checkForSnugget(shapefile, section, order, filterColumn, filterVal)
      if oldSnugget is not None and oldSnugget not in oldSnuggets:
        oldSnuggets.append(oldSnugget)
      overwriteAll = askUserAboutOverwriting(row, oldSnuggets, overwriteAll)
      processSnugget(row, shapefile, section, order, filterColumn, filterVal)

  return overwriteAll


def processSnugget(row, shapefile, section, order, filterColumn, filterVal):
  removeOldSnugget(shapefile, section, order, filterColumn, filterVal)
  if row["image_slideshow_folder"] is not '':
    addSlideshowSnugget(row, shapefile, section, filterColumn, filterVal)
  elif row["video"] is not '':
    addVideoSnugget(row, shapefile, section, filterColumn, filterVal)
  else:
    addTextSnugget(row, shapefile, section, filterColumn, filterVal)



def getShapefileClass(row):
  modelName = row["shapefile"].lower()
  if ContentType.objects.filter(app_label=appName, model=modelName).exists():
    shapefile = ContentType.objects.get(app_label=appName, model=modelName).model_class()
    return shapefile
  else:  # this means that nothing was found in the database for the shapefile name we read from snuggetFile
    print("No shapefile with the name", row["shapefile"], "appears to have been loaded.")
    print("If the shapefile exists, you may still need to run the migration and loading steps - see the 'Load some data' section of the readme file.")
    return None


def getFilterFieldName(shapefile):
  # The lookup value / filter field is the one from the shapefile that is not one of these.
  field = next(f for f in shapefile._meta.get_fields() if f.name not in ['id', 'geom', 'group'])
  return field.name


def getShapefileFilter(shapefile, filterVal):
  fieldName = getFilterFieldName(shapefile)
  if fieldName == 'rast':
    for tile in shapefile.objects.defer("rast").all():
      if int(filterVal) in tile.rast.bands[0].data():
        return int(filterVal)
    print(
      "Lookup value of", filterVal,
      "not found in", shapefile,
      ", the first tile of which only has values between", str(shapefile.objects.first().rast.bands[0].min),
      "and", str(shapefile.objects.first().rast.bands[0].max), "inclusive."
    )
  else:
    kwargs = {fieldName: filterVal}
    if shapefile.objects.filter(**kwargs).exists():
      return shapefile.objects.get(**kwargs)
    else:
      print("Could not find a filter field for", shapefile)
  return None


def addPopOutIfExists(row):
  if (row["pop_out_image"] != '' or row["pop_out_txt"] != '' or row["pop_alt_txt"] != '' or row["pop_out_link"] != '' or row["pop_out_video"] is not None):
    print('Creating pop-out section with values ', row["pop_out_image"], row["pop_out_txt"], row["pop_alt_txt"], row["pop_out_link"], row['pop_out_video'] )
    popout = SnuggetPopOut.objects.create(text=row["pop_out_txt"], alt_text=row["pop_alt_txt"], link=row["pop_out_link"], video=row["pop_out_video"])
    if row["pop_out_image"] != '':
      imageFile = os.path.join(dataDir, 'images/pop_out', row["pop_out_image"])
      with open(imageFile, 'rb') as f:
        data = File(f)
        popout.image.save(row["pop_out_image"], data, True)

    return popout
  else:
    return None



def addTextSnugget(row, shapefile, section, filterColumn, filterVal):
  group = shapefile.getGroup()
  shapefileFilter = getShapefileFilter(shapefile, filterVal)

  if shapefileFilter is not None:
    kwargs = {
      'section': section,
      'group': group,
      filterColumn: shapefileFilter,
      'content': row["text"],
      'percentage': row["intensity"],
      'order': row['txt_location']
    }
    snugget = TextSnugget.objects.create(**kwargs)
    snugget.pop_out = addPopOutIfExists(row)
    snugget.save()
    print('Created snugget:', snugget)



def addVideoSnugget(row, shapefile, section, filterColumn, filterVal):
  group = shapefile.getGroup()
  shapefileFilter = getShapefileFilter(shapefile, filterVal)

  kwargs = {
    'section': section,
    'group': group,
    filterColumn: shapefileFilter,
    'text': row["text"],
    'video': row["video"],
    'percentage': row["intensity"],
    'order': row['txt_location']
  }
  snugget = EmbedSnugget.objects.create(**kwargs)
  snugget.save()
  print('Created embed snugget:', snugget)



def addSlideshowSnugget(row, shapefile, section, filterColumn, filterVal):
  group = shapefile.getGroup()
  shapefileFilter = getShapefileFilter(shapefile, filterVal)

  kwargs = {
    'section': section,
    'group': group,
    filterColumn: shapefileFilter,
    'text': row["text"],
    'percentage': row["intensity"],
    'order': row['txt_location']
  }
  snugget = SlideshowSnugget.objects.create(**kwargs)
  snugget.pop_out = addPopOutIfExists(row)
  addSlideshow(os.path.join(dataDir, 'images/', row["image_slideshow_folder"]), snugget)
  snugget.save()
  print('Created slideshow snugget:', snugget)


def addSlideshow(folder, snugget):
  slideshowFile = os.path.join(folder, slideshowFilename)

  slides = XLSXDictReader(slideshowFile)
  rowCount = 1 # row 1 consists of field names, so row 2 is the first data row. We'll increment this before first referencing it.
  for row in slides:
    rowCount += 1
    photo = PastEventsPhoto.objects.create(snugget=snugget, caption=row["caption"])
    if row["image"] != '':
      imageFile = os.path.join(folder, row["image"])
      with open(imageFile, 'rb') as f:
        data = File(f)
        photo.image.save(row["image"], data, True)
    print("...... Created", photo)

  print("... Image slideshow created from", folder)


def findAllFilterVals(shapefile):
  fieldName = getFilterFieldName(shapefile)
  if fieldName == 'rast':
    # If it's a raster, we take a small shortcut and just return the whole
    # range of values from min to max, assuming all the intermediate ones exist.
    # This may create some surplus snuggets, but won't miss any that should be there.
    minima = [256]
    maxima = [0]
    for tile in shapefile.objects.all():
      if tile.rast.bands[0].min is not None:
        minima.append(tile.rast.bands[0].min)
        maxima.append(tile.rast.bands[0].max)
    return list(range(min(minima), max(maxima)+1))
  else:
    return shapefile.objects.values_list(fieldName, flat=True)


def checkForSnugget(shapefile, section, order, filterColumn, filterVal):
  filter = getShapefileFilter(shapefile, filterVal)
  if filter is not None:
    kwargs = {'section': section, 'order': order, filterColumn: filter}
    if Snugget.objects.filter(**kwargs).exists():
      return Snugget.objects.select_subclasses().get(**kwargs)
  return None


def removeOldSnugget(shapefile, section, order, filterColumn, filterVal):
  kwargs = {'section': section, filterColumn: getShapefileFilter(shapefile, filterVal), 'order': order}
  Snugget.objects.filter(**kwargs).delete()


# Check with the user about overwriting existing snuggets, giving them the options to:
# either quit and check what's going on, or say "yes to all" and not get prompted again.
def askUserAboutOverwriting(row, old, overwriteAll):
  if overwriteAll: # if it's already set, then don't do anything else
    return True
  else:
    if old is not None and not isinstance(old, list):
      print("In shapefile ", repr(row["shapefile"]), " there is already a snugget defined for section " , repr(row["section"]), ", intensity ", repr(row["lookup_value"]), ", txt_location ", repr(row["txt_location"]), " with the following text content:", sep="")
      print(old)
    elif old and isinstance(old, list):
      print("In shapefile ", repr(row["shapefile"]), " there are existing snuggets for section" , repr(row["section"]), ", txt_location", repr(row["txt_location"]), " with the following text content:", sep="")
      for snugget in old:
        print(snugget)
    else:
      # if no existing snuggets were found, then we neither need to ask the user this time nor change overwriteAll
      return overwriteAll

    print("Please enter one of the following letters to choose how to proceed:")
    print("R: Replace the existing snugget[s] with the new value loaded from", snuggetFile, " and ask again for the next one.")
    print("A: replace All without asking again.")
    print("Q: quit so you can edit", snuggetFile, "and/or check the snuggets in the Django admin panel and try again.")
    response = ""
    while response not in ["A", "R", "Q"]:
      response = input(">> ").upper()

    if response == "Q":
      exit(0)
    elif response == "A":
      return True
    elif response == "R":
      return False



# drop-in replacement for built-in csv.DictReader() function with .xlsx files
# originally from https://gist.github.com/mdellavo/853413
# then heavily adapted first to make it work, then to simplify, and finally with suggestions from later commenters on that gist
def XLSXDictReader(fileName, sheetName=None):
  book = openpyxl.reader.excel.load_workbook(fileName)
  # if there's no sheet name specified, try to get the active sheet.  This will work reliably for workbooks with only one sheet; unpredictably if there are multiple worksheets present.
  if sheetName is None:
    sheet = book.active
  elif sheetName not in book.sheetnames:
    print(sheetName, "not found in", fileName)
    exit()
  else:
    sheet = book[sheetName]

  rows = sheet.max_row + 1
  cols = sheet.max_column + 1

  def cleanValue(s):
    if s == None:
      return ''
    else:
      return str(s).strip()

  def item(i, j):
    return (
      cleanValue(sheet.cell(row=1, column=j).value),
      cleanValue(sheet.cell(row=i, column=j).value)
    )

  return (dict(item(i,j) for j in range(1, cols)) for i in range(2, rows))
